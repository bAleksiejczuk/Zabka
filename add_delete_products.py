import tkinter as tk
from tkinter import messagebox
import pandas as pd
import os
from datetime import datetime
import pytz
from openpyxl import load_workbook
import re

def add_product(name_entry, count_entry, price_entry):
    name = name_entry.get()
    count = count_entry.get()
    price = price_entry.get()

    if name and count and price:
        try:
            if len(name) > 20:
                messagebox.showerror("Błąd", "Nazwa produktu może mieć maksymalnie 20 znaków!")
                return

            count = int(count)
            if count <= 0:
                messagebox.showerror("Błąd", "Ilość musi być większa niż 0!")
                return
            if count > 99999:
                messagebox.showerror("Błąd", "Ilość nie może przekraczać 99999!")
                return

            
            if not re.match(r'^\d+\.\d{2}$', price):
                messagebox.showerror("Błąd", "Cena musi być w formacie XX.XX (np. 12.99)!")
                return

            price_value = float(price)
            if price_value <= 0:
                messagebox.showerror("Błąd", "Cena musi być większa niż 0!")
                return
            if price_value > 99999.99:
                messagebox.showerror("Błąd", "Cena nie może przekraczać 99999.99!")
                return

            file_path = os.path.join("data", "products.xlsx")

            warsaw_tz = pytz.timezone("Europe/Warsaw")
            now_warsaw = datetime.now(warsaw_tz)
            formatted_date = now_warsaw.strftime("%Y-%m-%d")

            # Wczytaj plik i nazwę pierwszego arkusza
            excel_file = pd.ExcelFile(file_path)
            sheet_name = excel_file.sheet_names[0]
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, skiprows=1)

            new_id = 1  # Domyślne ID dla pierwszego produktu
            
            if not df.empty and len(df) > 0:
                try:
                    # Pobierz wszystkie ID z pierwszej kolumny
                    ids = []
                    for index, row in df.iterrows():
                        row_data = str(row[0]).split(',')
                        if len(row_data) > 0 and row_data[0].isdigit():
                            ids.append(int(row_data[0]))
                    
                    if ids:
                        new_id = max(ids) + 1
                except:
                    new_id = 1  

            # Nowy wiersz do dodania
            new_row = f"{new_id},{name},{price},{count},{formatted_date}"
            new_df = pd.DataFrame([[new_row]])

            # Załaduj workbook i arkusz
            book = load_workbook(file_path)
            sheet = book[sheet_name]
            next_row = sheet.max_row + 1

            # Wstaw dane
            for row_index, row in enumerate(new_df.values, start=next_row):
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=col_index, value=value)

            book.save(file_path)

            messagebox.showinfo("Dodano", f"Dodano produkt: {name} (ID: {new_id}, Ilość: {count}, Cena: {price} zł)")

            name_entry.delete(0, tk.END)
            count_entry.delete(0, tk.END)
            price_entry.delete(0, tk.END)

        except ValueError as ve:
            messagebox.showerror("Błąd", f"Błąd walidacji danych: {str(ve)}")
        except Exception as e:
            messagebox.showerror("Błąd", f"Wystąpił problem przy dodawaniu produktu:\n{str(e)}")
    else:
        messagebox.showwarning("Brak danych", "Wprowadź nazwę, ilość i cenę produktu.")




def delete_product(delete_entry):
    name_for_delete = delete_entry.get()
    if not name_for_delete:
        messagebox.showwarning("Brak danych", "Wprowadź nazwę produktu do usunięcia.")
        return

    try:
        file_path = os.path.join("data", "products.xlsx")
        df_raw = pd.read_excel(file_path, header=None)

        df_split = df_raw[0].str.split(',', expand=True)

        # Ustaw nagłówki z pierwszego wiersza
        df_split.columns = df_split.iloc[0]
        df_split = df_split[1:]  # usuń pierwszy wiersz z nagłówkami

        # Znajdź wiersze dopasowane do podanej nazwy
        matches = df_split['PRODUCT'].str.strip().str.lower() == name_for_delete.strip().lower()
        if not matches.any():
            messagebox.showinfo("Brak produktu", f"Nie znaleziono produktu o nazwie: {name_for_delete}")
            return

        
        deleted_products = df_split[matches]
        deleted_info = []
        for _, row in deleted_products.iterrows():
            product_id = row.get('ID', 'Brak ID')
            product_name = row.get('PRODUCT', 'Brak nazwy')
            product_quantity = row.get('NO_PACKAGES_AVAILABLE', 'Brak ilości')
            product_price = row.get('PRICE', 'Brak ceny')
            deleted_info.append(f"ID: {product_id}, Nazwa: {product_name}, Ilość: {product_quantity}, Cena: {product_price} zł")

        # Usuń wiersze dopasowane do podanej nazwy
        df_split = df_split[~matches]

        # Połącz dane z powrotem w jedną kolumnę A
        df_combined = df_split.apply(lambda row: ','.join(row.dropna().astype(str)), axis=1)
        df_combined = pd.DataFrame(df_combined)
        
        # Dodaj nagłówki z powrotem jako pierwszy wiersz
        header_row = ','.join(df_split.columns)
        df_combined.loc[-1] = header_row  # dodaj jako pierwszy wiersz
        df_combined.index = df_combined.index + 1
        df_combined = df_combined.sort_index()

        # Zapisz z powrotem do pliku
        df_combined.to_excel(file_path, index=False, header=False)

        # Wyświetl szczegółowy komunikat z informacjami o usuniętych produktach
        if len(deleted_info) == 1:
            messagebox.showinfo("Usunięto", f"Usunięto produkt:\n{deleted_info[0]}")
        else:
            messagebox.showinfo("Usunięto", f"Usunięto {len(deleted_info)} produktów:\n" + "\n".join(deleted_info))
        
        delete_entry.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Błąd", f"Wystąpił błąd podczas usuwania produktu:\n{e}")

